from datetime import datetime, timedelta

import pandas as pd
import reversion
from apps.HWCheck.models import Document
from apps.measurement.models import MeasurementRequest
from django.core.mail import EmailMessage
from django.db import connection
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from rest_framework import exceptions, status, viewsets
from rest_framework.response import Response


def remove_creator(classA):
    fields_list = [field.name for field in classA._meta.get_fields()]
    fields_list.remove("creator")
    return tuple(fields_list)


def bulk_create(self, request):
    if isinstance(request.data, list):
        for instance in request.data:
            serializer = self.get_serializer(data=instance)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
    else:
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
    return serializer


def synchronize_attributes(A, B, synchron_list, additional, diff):
    for attr in synchron_list:
        if attr == "common_user":
            B.common_user.set(A.common_user.all())
        elif getattr(B, attr) != getattr(A, attr):
            diff[attr] = str(getattr(B, attr)) + " -> " + str(getattr(A, attr))
            setattr(B, attr, getattr(A, attr))

    if additional:
        for attr in additional:
            diff[attr] = str(getattr(B, attr)) + " -> "
            setattr(B, attr, additional[attr])
            diff[attr] += str(getattr(B, attr))

    B.save()


def synchronizer_B_from_A(
    synchron_list, A, B, additional={}, dereference_id="", create=False
):
    diff = {}

    if create:
        if isinstance(A, str):
            A = globals()[A]()  # 문자열로 주어진 이름으로 객체 생성
        if isinstance(B, str):
            B = globals()[B]()  # 문자열로 주어진 이름으로 객체 생성
        synchronize_attributes(A, B, synchron_list, additional, diff)
    else:
        if isinstance(B, Document):
            with reversion.create_revision():
                synchronize_attributes(A, B, synchron_list, additional, diff)
                reversion.set_comment(diff)
        else:
            synchronize_attributes(A, B, synchron_list, additional, diff)
    if dereference_id:
        setattr(A, dereference_id, B.id)
        A.save()

    return diff, A, B


def create_empty_excel():
    # 빈 DataFrame 생성
    df = pd.DataFrame()

    # 빈 엑셀 파일 생성
    writer = pd.ExcelWriter("example_data/empty_excel.xlsx")
    df.to_excel(writer, sheet_name="Sheet1", index=False)
    writer.book.save("example_data/empty_excel.xlsx")


def make_excel(queryset_db):
    create_empty_excel()
    writer = pd.ExcelWriter("example_data/empty_excel.xlsx")
    for db in queryset_db:
        table_cols = ", ".join(db["db_columns"])
        if len(db["data"]) == 1:
            db_data = "('" + db["data"][0] + "')"
        else:
            db_data = tuple(db["data"])
        sql = f"SELECT {table_cols} FROM {db['db_table']} WHERE {db['db_columns'][0]} IN {db_data};"
        dataframe = pd.DataFrame(
            get_objects_from_table(db["db_table"], sql)
        ).reset_index(drop=True)
        dataframe.columns = db["db_columns"]
        sheet_name = db["title"]  # 시트 이름은 db.title로 설정
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.book.save("example_data/data.xlsx")
    return "example_data/data.xlsx"


def get_objects_from_table(table_name, sql):
    with connection.cursor() as cursor:
        cursor.execute(sql)
        results = cursor.fetchall()

    return results


def make_comment(string):

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="comment",
            fontName="Helvetica-Bold",  # "Helvetica-Bold" 폰트로 굵게 설정
            fontSize=24,  # 폰트 크기 24로 설정 (2배 크기)
            alignment=1,  # 가운데 정렬
        )
    )
    hangul1 = string
    spacer = Spacer(1, 36)
    return Paragraph(hangul1, style=styles["comment"]), spacer


def listing(self, request, queryset, *args, **kwargs):
    column_name = self.kwargs("column_name")
    queryset = queryset.values_list(column_name, flat=True)
    return Response(queryset, status=status.HTTP_204_NO_CONTENT)


def convert_excel_to_pdf(excel_file, pdf_file):
    # 엑셀 파일 로드
    wb = load_workbook(excel_file, read_only=True)
    # 글꼴 적용
    pdfmetrics.registerFont(TTFont("NanumGothic", "static/fonts/NanumGothic.ttf"))
    # PDF 생성
    doc = SimpleDocTemplate(pdf_file, pagesize=letter)
    elements = []

    # 테이블 스타일 설정
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "NanumGothic"),
            ("FONTSIZE", (0, 0), (-1, 0), 16),  # 폰트 크기 추가
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )

    for sheet_name in wb.sheetnames:
        # 데이터프레임으로 변환
        ws = wb[sheet_name]
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        # 책갈피 추가
        if elements:
            elements.append(PageBreak())
        data_style = TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "NanumGothic"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
            ]
        )
        # 데이터프레임을 테이블로 변환
        df_table = Table([df.columns.tolist()] + df.values.tolist(), repeatRows=1)
        df_table.setStyle(style)
        df_table.setStyle(data_style)

        elements.append(Paragraph(sheet_name, getSampleStyleSheet()["Heading1"]))
        elements.append(df_table)

    doc.build(elements)


def default_response(data, status=status.HTTP_204_NO_CONTENT):
    if data or data == []:
        response_data = {
            "datetime": timezone.now(),
            "message": "OK",
            "data": data,
        }
    else:
        response_data = {
            "datetime": timezone.now(),
            "message": "OK",
        }

    return Response(response_data, status)
